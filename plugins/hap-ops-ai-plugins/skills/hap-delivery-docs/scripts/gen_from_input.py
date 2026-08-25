#!/usr/bin/env python3
# 读「入参模板」xlsx → 生成该项目全套真实交付文档（8 类）。
# 占位替换：192.168.1.x→实际IP、hap.domain.com→实际域名、某示例客户→实际客户名、实施目录占位→实际路径；
# 凭据实际值回填《凭据登记表》对应行；架构图 SVG 替换后重渲 PNG。
# 密码不写入部署/运维等文档（保留 <强密码> 占位，按惯例仅在《凭据登记表》加密留存）。
# 用法: python gen_from_input.py <入参.xlsx> <outdir>
import os, sys, subprocess, zipfile, re
import openpyxl
try: import fitz
except: fitz=None
_HERE=os.path.dirname(os.path.abspath(__file__))
INP=sys.argv[1]; OUT=sys.argv[2] if len(sys.argv)>2 else 'delivery_out'
os.makedirs(OUT,exist_ok=True)
PY=sys.executable

# ---------- 1. 读入参 ----------
wb=openpyxl.load_workbook(INP,data_only=True)
def read_last_sheet_notes():
    """Read the final sheet before generation; it carries project-specific constraints."""
    ws=wb[wb.sheetnames[-1]]
    notes=[]
    for row in ws.iter_rows(values_only=True):
        vals=[str(v).strip() for v in row if v is not None and str(v).strip()]
        if vals:
            notes.append(" | ".join(vals))
    return ws.title, notes

custom_sheet_name, custom_notes = read_last_sheet_notes()
if custom_notes:
    print("生成前自定义项描述（来自最后一个 Sheet：%s）:"%custom_sheet_name)
    for line in custom_notes:
        print("  - "+line)

def sheet(key):
    for s in wb.sheetnames:
        if key in s: return wb[s]
    raise SystemExit("入参缺少 Sheet: "+key)
def kv(ws, kcol=1, vcol=4, start=4):
    d={}
    for r in range(start, ws.max_row+1):
        k=ws.cell(r,kcol).value
        if k: d[str(k).strip()]=(ws.cell(r,vcol).value or "")
    return d

pj=kv(sheet("项目信息"),1,4)
def g(*names):
    for n in names:
        for k,v in pj.items():
            if n in k and str(v).strip(): return str(v).strip()
    return ""
customer=g("客户名称"); verraw=g("部署版本"); sceneraw=g("部署场景")
brand=g("品牌") or "mingdao"; domain=g("主访问域名"); date=g("交付日期") or "2026/06/23"
impl=g("实施目录")
VER={"集群精简版":"lite","集群标准版":"std","集群专业版":"pro"}.get(verraw,"")
if not VER: raise SystemExit("②项目信息·部署版本 必填（集群精简版/标准版/专业版），当前=%r"%verraw)
SC={"场景A":"A","场景B":"B"}.get(sceneraw,"")
if VER!="lite" and not SC: raise SystemExit("标准/专业版需选场景A或B，当前=%r"%sceneraw)

# IP 映射(占位→实际)
ipmap={}
wsip=sheet("节点IP")
for r in range(4, wsip.max_row+1):
    ph=wsip.cell(r,1).value; real=wsip.cell(r,4).value
    if ph and real and str(real).strip(): ipmap[str(ph).strip()]=str(real).strip()
# 凭据
creds=kv(sheet("凭据清单"),1,5)

# ---------- 2. 跑 8 类生成 ----------
def run(script,*args):
    env=dict(os.environ,PYTHONUTF8="1",PYTHONIOENCODING="utf-8",COVER_DATE=date)
    r=subprocess.run([PY,os.path.join(_HERE,script),*[str(a) for a in args]],env=env,
                     capture_output=True,text=True,encoding="utf-8")
    if r.returncode!=0: raise SystemExit("%s 失败:\n%s"%(script,(r.stderr or r.stdout)))
    return r.stdout
sc=[SC] if SC else []
print("生成中（版本=%s 场景=%s）..."%(VER, SC or "无"))
run("gen_deploy.py", VER, *sc, OUT)
run("gen_ops.py",    VER, *sc, OUT)
run("gen_mig.py",    VER, OUT)
run("gen_chk.py",    VER, *sc, OUT)
run("gen_ref.py",    "faq", OUT)
run("gen_ref.py",    "resource", OUT)
run("gen_arch.py",   VER, *sc, OUT)
run("gen_cred.py",   VER, *sc, OUT)

# ---------- 3. 占位替换 map（仅唯一安全 token；密码不入文档）----------
rep=[]
for ph,real in sorted(ipmap.items(), key=lambda x:-len(x[0])): rep.append((ph,real))
if domain: rep.append(("hap.domain.com",domain))
if customer: rep.append(("某示例客户",customer))
if impl and impl!="<交付方现场实施目录>": rep.append(("<交付方现场实施目录>",impl))
def apply_text(t):
    for a,b in rep: t=t.replace(a,b)
    return t

def repl_zip(path):   # docx/xlsx：zip 内所有 xml 文本替换
    tmp=path+".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED) as zo:
        for it in zin.infolist():
            data=zin.read(it.filename)
            if it.filename.endswith((".xml",".rels")):
                try: data=apply_text(data.decode("utf-8")).encode("utf-8")
                except: pass
            zo.writestr(it,data)
    os.replace(tmp,path)

# ---------- 4. 处理产物 ----------
files=sorted(os.listdir(OUT))
for f in files:
    p=os.path.join(OUT,f)
    if f.endswith(".docx"):
        repl_zip(p)
    elif f.endswith(".svg"):
        # 架构图节点为完整 IP，随入参全量替换（含 IP/域名）
        t=apply_text(open(p,encoding="utf-8").read())
        open(p,"w",encoding="utf-8").write(t)
        if fitz:
            png=p[:-4]+".png"
            pix=fitz.open(p)[0].get_pixmap(matrix=fitz.Matrix(2,2)); pix.save(png)   # 重渲 PNG

# 凭据登记表：替换占位 + 回填实际值
CRED_MATCH={
 "MySQL root 密码":["root 初始化密码","ENV_MYSQL_PASSWORD"],
 "MongoDB admin(root) 密码":["admin 管理账号"],
 "MongoDB 业务库(hap) 密码":["业务库账号"],
 "MongoDB 副本集 keyFile":["副本集 keyFile"],
 "Redis 密码":["requirepass","masterauth","Sentinel auth-pass","ENV_REDIS_SENTINEL_PASSWORD","ENV_REDIS_PASSWORD"],
 "Elasticsearch elastic 密码":["x-pack 密码","ENV_ELASTICSEARCH_PASSWORD"],
 "MinIO ROOT_USER":["MINIO_ROOT_USER"],
 "MinIO ROOT_PASSWORD":["MINIO_ROOT_PASSWORD"],
 "File AccessKey":["ENV_ACCESS_KEY_FILE"],
 "File SecretKey":["ENV_SECRET_KEY_FILE"],
 "ENV_API_TOKEN":["ENV_API_TOKEN"],
 "Keepalived auth_pass":["auth_pass"],
 "Grafana 登录密码":["Grafana 登录"],
 "HAP 平台超级管理员":["平台超级管理员"],
}
filled=0
for f in files:
    if "凭据登记表" not in f or not f.endswith(".xlsx"): continue
    p=os.path.join(OUT,f); w=openpyxl.load_workbook(p); ws=w["凭据登记表"]
    # 表头定位
    hdr=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]
    cKey=hdr.index("配置项 / 字段")+1; cVal=[i for i,h in enumerate(hdr) if h and "实际值" in str(h)][0]+1
    for r in range(5, ws.max_row+1):
        for c in range(1,ws.max_column+1):
            v=ws.cell(r,c).value
            if isinstance(v,str): ws.cell(r,c).value=apply_text(v)        # IP/域名/客户名
        key=str(ws.cell(r,cKey).value or "")
        for ck,subs in CRED_MATCH.items():
            if creds.get(ck) and any(s in key for s in subs):
                ws.cell(r,cVal).value=str(creds[ck]).strip(); filled+=1; break
    w.save(p)

print("完成 → %s"%OUT)
print("  版本=%s 场景=%s 客户=%s 域名=%s"%(VER,SC or "无",customer or "(未填)",domain or "(未填)"))
print("  IP 替换 %d 项 | 凭据回填 %d 处 | 产物 %d 个"%(len(ipmap),filled,len([x for x in files])))
miss=[k for k in ipmap if not ipmap[k]]
notfilled=[ck for ck in CRED_MATCH if not creds.get(ck)]
if notfilled: print("  未填凭据(保留占位):", "、".join(notfilled))
