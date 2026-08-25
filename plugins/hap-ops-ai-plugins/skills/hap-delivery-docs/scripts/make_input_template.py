# 生成 HAP 交付文档「入参模板」xlsx：用户每项目填一份，据此把固化模板占位替换为真实值生成交付件。
import openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
out=sys.argv[1] if len(sys.argv)>1 else 'HAP交付入参模板.xlsx'
F="Microsoft YaHei"
H=PatternFill("solid",fgColor="1E5BA8"); HF=Font(name=F,bold=True,color="FFFFFF",size=10.5)
Y=PatternFill("solid",fgColor="FFF6CC")          # 待填(黄)
G=PatternFill("solid",fgColor="EAF1FB")          # 分组(浅蓝)
cf=Font(name=F,size=10); tf=Font(name=F,bold=True,size=15,color="1E5BA8"); nf=Font(name=F,size=9,color="888888")
bf=Font(name=F,bold=True,size=10,color="0C447C")
thin=Side(style="thin",color="D0D7E5"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
C=Alignment(horizontal="center",vertical="center",wrap_text=True); L=Alignment(horizontal="left",vertical="center",wrap_text=True)
wb=openpyxl.Workbook()

def header(ws,cols,row=1):
    for j,c in enumerate(cols,1):
        x=ws.cell(row,j,c); x.fill=H; x.font=HF; x.alignment=C; x.border=bd
def put(ws,r,vals,fillcols=()):
    for j,v in enumerate(vals,1):
        x=ws.cell(r,j,v); x.font=cf; x.border=bd; x.alignment=L
        if (j) in fillcols: x.fill=Y
def grouprow(ws,r,text,ncol):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncol)
    x=ws.cell(r,1,text); x.fill=G; x.font=bf; x.alignment=L; x.border=bd

# ===== Sheet 0 填写说明 =====
ws=wb.active; ws.title="① 填写说明"
ws["A1"]="HAP 私有部署交付文档 · 入参模板"; ws["A1"].font=tf
guide=[
 "",
 "用途：每个项目整理这一份 Excel（填黄色「实际值」列），交回后即可据此生成全套真实交付文档；无需逐份改文档。",
 "",
 "填写顺序：",
 "  1) 「② 项目信息」：客户名、部署版本、场景、域名、日期等（版本/场景/品牌为下拉选择）。",
 "  2) 「③ 节点 IP 规划」：按你选的【部署版本】只填对应行的「实际 IP」。占位 IP（192.168.1.x）只是模板示例，按实际网络填。",
 "  3) 「④ 凭据清单」：填每项「实际值」。注意「一致性」列——同一密码用于多处时必须保持一致（部署/迁移时同样要求）。",
 "  4) 「⑤ 端口与默认值」：一般无需改（如确有变更再填覆盖值）。",
 "  5) 「⑥ 高级可选」：时区/子路径/双地址/证书等个性化项，用不到就留空。",
 "  6) 最后一个 Sheet「⑦ 自定义项描述」：填写本项目特殊约束；生成文档前必须先读取并据此调整。",
 "",
 "生成时的替换规则（你不用管，交回我处理）：",
 "  · 占位 IP 192.168.1.x → 你填的实际 IP；hap.domain.com → 实际域名；某示例客户 → 实际客户名。",
 "  · <强密码> / mingdao / storage 等 → 你填的实际凭据；凭据实际值进《凭据登记表》对应列并提示加密保管。",
 "",
 "重要：",
 "  · 只填你所选版本用到的行（精简版 6 节点 / 标准版 17 / 专业版 29）。多余行可留空。",
 "  · 场景仅标准/专业版区分 A（独立 docker swarm）/ B（同一 Swarm·Node01 编排）；精简版选「不适用」。",
 "  · 凭据为敏感信息：本表填好后请加密保管，勿随文档明文外发。",
]
r=2
for g in guide:
    ws.cell(r,1,g).font=(bf if g.endswith("：") and not g.startswith("  ") else nf if g.startswith("  ") or g=="" else cf)
    r+=1
ws.column_dimensions["A"].width=120

# ===== Sheet ② 项目信息 =====
ws=wb.create_sheet("② 项目信息")
ws["A1"]="项目信息（黄色为必填 / 选填的实际值）"; ws["A1"].font=tf
header(ws,["参数","说明","示例 / 占位","实际值（填写）"],3)
proj=[
 ("客户名称","出现在交付清单正文与封面","某示例客户"),
 ("部署版本","三选一（下拉）","集群专业版"),
 ("部署场景","仅标准/专业版区分（下拉）；精简版选 不适用","场景A"),
 ("品牌","域名体系（下拉）：mingdao 或 nocoly","mingdao"),
 ("主访问域名","HAP 前端访问域名（含或不含 https 由证书定）","hap.domain.com"),
 ("是否有上游 LB/网关","有则填其 IP；无则留空（精简版通常无）","192.168.1.10"),
 ("交付日期","封面 / 清单日期，格式 YYYY/MM/DD","2026/06/23"),
 ("项目名称","封面项目名","HAP 超级应用平台 · 私有部署项目"),
 ("交付方","封面 / 签字","明道云"),
 ("项目对接人","清单签字","（填写）"),
 ("客户接收人","清单签字","（填写）"),
 ("交付件实施目录","交付件在交付方现场的存放目录","<交付方现场实施目录>"),
]
r=4
for (k,d,ex) in proj:
    put(ws,r,[k,d,ex,""],fillcols=(4,)); r+=1
for j,w in enumerate([22,40,28,30],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"
# 下拉
dv_ver=DataValidation(type="list",formula1='"集群精简版,集群标准版,集群专业版"',allow_blank=True); ws.add_data_validation(dv_ver); dv_ver.add("D5")
dv_sc=DataValidation(type="list",formula1='"场景A,场景B,不适用（精简版）"',allow_blank=True); ws.add_data_validation(dv_sc); dv_sc.add("D6")
dv_br=DataValidation(type="list",formula1='"mingdao,nocoly"',allow_blank=True); ws.add_data_validation(dv_br); dv_br.add("D7")

# ===== Sheet ③ 节点 IP 规划 =====
ws=wb.create_sheet("③ 节点IP规划")
ws["A1"]="节点 IP 规划（按所选【部署版本】填对应行的「实际 IP」；占位 IP 仅为模板示例）"; ws["A1"].font=tf
header(ws,["占位 IP","角色 / 组件","适用版本","实际 IP（填写）"],3)
ips=[
 ("192.168.1.10","上游 LB / 网关（可选）","通用"),
 ("192.168.1.20","Nginx VIP（Keepalived）/ 精简版为 Nginx 单节点","通用"),
 ("192.168.1.11","Nginx 01（MASTER）","标准 / 专业"),
 ("192.168.1.12","Nginx 02（BACKUP）","标准 / 专业"),
 ("192.168.1.21","K8s Master 01 / 精简版 K8s 主","通用"),
 ("192.168.1.22","K8s Master 02 / 精简版 K8s 从","通用"),
 ("192.168.1.23","K8s Master 03","标准 / 专业"),
 ("192.168.1.24","K8s Worker 01","专业"),
 ("192.168.1.25","K8s Worker 02","专业"),
 ("192.168.1.31","MySQL 01（专业）/ 数据库共置 01（标准）/ 单机数据库（精简：MySQL+Mongo+Redis）","通用"),
 ("192.168.1.32","MySQL 02（专业）/ 数据库共置 02（标准）","标准 / 专业"),
 ("192.168.1.33","MySQL 03（专业）/ 数据库共置 03（标准）","标准 / 专业"),
 ("192.168.1.34","MongoDB 01（专业，独立副本集）","专业"),
 ("192.168.1.35","MongoDB 02（专业）","专业"),
 ("192.168.1.36","MongoDB 03（专业）","专业"),
 ("192.168.1.41","Redis 01（Master + Sentinel）","标准 / 专业"),
 ("192.168.1.42","Redis 02","标准 / 专业"),
 ("192.168.1.43","Redis 03","标准 / 专业"),
 ("192.168.1.51","Kafka 01（专业）/ 中间件共置 01（标准）/ 单机中间件（精简：Kafka+ES+MinIO+File）","通用"),
 ("192.168.1.52","Kafka 02（专业）/ 中间件共置 02（标准）","标准 / 专业"),
 ("192.168.1.53","Kafka 03（专业）/ 中间件共置 03（标准）","标准 / 专业"),
 ("192.168.1.54","中间件共置 04（标准）","标准"),
 ("192.168.1.61","ES 01（专业）/ Flink 01（标准）","标准 / 专业"),
 ("192.168.1.62","ES 02（专业）/ Flink 02（标准）","标准 / 专业"),
 ("192.168.1.63","ES 03（专业）","专业"),
 ("192.168.1.71","存储 01 · MinIO+File（专业）","专业"),
 ("192.168.1.72","存储 02 · MinIO+File（专业）","专业"),
 ("192.168.1.73","存储 03 · MinIO+File（专业）","专业"),
 ("192.168.1.74","存储 04 · MinIO+File（专业）","专业"),
 ("192.168.1.81","Flink 01（专业，独立）","专业"),
 ("192.168.1.82","Flink 02（专业）","专业"),
 ("192.168.1.83","Flink 03（专业）","专业"),
 ("192.168.1.30","Flink（精简版单节点）","精简"),
]
r=4
for (ip,role,ver) in ips:
    put(ws,r,[ip,role,ver,""],fillcols=(4,)); r+=1
for j,w in enumerate([16,52,14,22],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

# ===== Sheet ④ 凭据清单 =====
ws=wb.create_sheet("④ 凭据清单")
ws["A1"]="凭据清单（填「实际值」；注意一致性列——同一值用于多处必须相同）"; ws["A1"].font=tf
header(ws,["凭据项","用途 / 配置位置","占位 / 默认","一致性要求","实际值（填写）"],3)
creds=[
 ("MySQL root 密码","MySQL 初始化 + ConfigMap ENV_MYSQL_PASSWORD","<强密码>","两处一致（约束 C）"),
 ("MongoDB admin(root) 密码","mongod 管理账号","<强密码>","—"),
 ("MongoDB 业务库(hap) 密码","mongod 业务账号 + ENV_MONGODB_URI","<强密码>","两处一致"),
 ("MongoDB 副本集 keyFile","/data/mongodb/keyfile（标准/专业）","<keyFile 内容>","副本集 3 节点完全一致（约束 F）；精简版不需要"),
 ("Redis 密码","redis.conf requirepass/masterauth + sentinel + ENV","<强密码>","集群版四处+3节点全一致（约束 B）；精简版 requirepass=ENV"),
 ("Elasticsearch elastic 密码","elasticsearch x-pack + ENV_ELASTICSEARCH_PASSWORD","<强密码>","两处一致（约束 D）"),
 ("MinIO ROOT_USER","minio.yaml（= HAP s3 / Flink s3 access-key）","mingdao","三处一致（约束 A）"),
 ("MinIO ROOT_PASSWORD","minio.yaml（= HAP s3 / Flink s3 secret-key）","<强密码>","三处一致（约束 A）"),
 ("File AccessKey","file ENV_ACCESS_KEY_FILE + ConfigMap ENV_FILE_ACCESSKEY","storage","两处一致（约束 E）"),
 ("File SecretKey","file ENV_SECRET_KEY_FILE + ConfigMap ENV_FILE_SECRETKEY","<强密码>","两处一致（约束 E）"),
 ("ENV_API_TOKEN","HAP ConfigMap（微服务鉴权）","<高熵随机字符串>","所有微服务副本一致（约束 G）"),
 ("Keepalived auth_pass","keepalived.conf（标准/专业）","HAP-Nginx-Keepalived-Auth","两 Nginx 节点一致（约束 H）；精简版不需要"),
 ("Grafana 登录密码","监控 Grafana admin","<强密码>","首次登录改密"),
 ("HAP 平台超级管理员","首次 Web 初始化创建","<账号> / <密码>","系统最高权限，妥善保管"),
]
r=4
for (k,d,ph,cons) in creds:
    put(ws,r,[k,d,ph,cons,""],fillcols=(5,)); r+=1
for j,w in enumerate([24,42,24,40,26],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

# ===== Sheet ⑤ 端口与默认值（参考） =====
ws=wb.create_sheet("⑤ 端口与默认值")
ws["A1"]="端口与默认值（一般无需修改；如确有变更填「覆盖值」，否则留空用默认）"; ws["A1"].font=tf
header(ws,["组件","默认端口 / 值","覆盖值（选填）"],3)
ports=[
 ("MinIO","9011-9014（精简版单节点 9011）"),("File 服务","9001-9004（精简版 9000）"),
 ("MySQL","3306 / Router 6446（精简版直连 3306）"),("MongoDB","27017"),
 ("Redis","6379 / Sentinel 26379"),("Kafka / ZooKeeper","9092 / 2181"),
 ("Elasticsearch","9200 / 9300"),("K8s / 安装管理器","6443（kube-apiserver）/ 38880（安装管理器/ENV_CAPTAIN_ENDPOINT）/ 38881（管理入口）"),
 ("HAP 微服务 www","8880（www 主地址）/ 18880（www 扩展地址，按需启用）"),
 ("Flink JobManager UI","28081"),("Nginx","80 / 443"),
 ("时区","Asia/Shanghai（ENV_TIME_ZONE）"),
]
r=4
for (k,v) in ports:
    put(ws,r,[k,v,""],fillcols=(3,)); r+=1
for j,w in enumerate([24,40,24],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

# ===== Sheet ⑥ 高级可选 =====
ws=wb.create_sheet("⑥ 高级可选")
ws["A1"]="高级 / 个性化（用不到留空）"; ws["A1"].font=tf
header(ws,["项","说明","示例","实际值（选填）"],3)
adv=[
 ("ENV_ADDRESS_MAIN","主访问地址，一般 = 主域名（自动）；主地址默认对应微服务端口 8880","https://hap.domain.com"),
 ("ENV_ADDRESS_ALLOWLIST","访问来源白名单（留空不限制）","内网网段 / 指定 IP"),
 ("ENV_MINGDAO_SUBPATH","子路径方式部署（如 /hap），不用留空","/hap"),
 ("双访问地址","对外第二地址 ENV_EXT_*（端口 18880）","hap1.domain.com"),
 ("SSL 证书","证书方式：上游网关终止 / Nginx 终止 / 无","上游网关终止"),
 ("NTP 时间源","时间同步服务器","ntp.aliyun.com"),
 ("ENV_TIME_ZONE","容器时区","Asia/Shanghai"),
 ("镜像品牌名","v7.1.0 起为 mingdaoyun-hap（一般自动）","mingdaoyun-hap"),
 ("其他个性化","你注意到的其他需要变更项，自由补充",""),
]
r=4
for (k,d,ex) in adv:
    put(ws,r,[k,d,ex,""],fillcols=(4,)); r+=1
for j,w in enumerate([24,40,28,28],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

# ===== Sheet ⑦ 自定义项描述（必须最后读取） =====
ws=wb.create_sheet("⑦ 自定义项描述")
ws["A1"]="自定义项描述（生成文档前必须先读取本 Sheet，并按描述调整）"; ws["A1"].font=tf
header(ws,["项目","描述 / 约束","影响范围","处理状态（生成前填写）"],3)
custom=[
 ("微服务端口口径","www 主地址默认端口为 8880；www 扩展地址如启用，默认端口为 18880；安装管理器/ENV_CAPTAIN_ENDPOINT 保持 38880，不得把 38880 当作 www 默认服务端口。","部署文档 / 运维文档 / 架构图 / 交付清单",""),
 ("项目特殊约束","如入口软件、入口数量、systemd 管理方式、主/扩展地址边界、组件部署节点、版本号、安装目录或附件配置等，请在此逐条填写。","全套交付件",""),
 ("生成前确认","Codex/生成脚本应先读取本 Sheet，形成项目约束清单；涉及文档内容、架构图、端口、systemd、地址、文件名或清单的，需在生成前或生成后专项修正并复核。","生成流程",""),
]
r=4
for (k,d,scope,status) in custom:
    put(ws,r,[k,d,scope,status],fillcols=(2,4)); r+=1
for j,w in enumerate([24,70,36,24],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

wb.save(out)
print("生成:",out)
for s in wb.sheetnames: print("  -",s)
