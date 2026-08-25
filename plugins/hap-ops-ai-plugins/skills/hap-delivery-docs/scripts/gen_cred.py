# 凭据登记表(基于固化部署基线,参数化 5 份)。用法: gen_cred.py <std|pro|lite> [A|B] <outdir>
import openpyxl, sys, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
ver=sys.argv[1] if len(sys.argv)>1 else 'pro'
scene=(sys.argv[2] if len(sys.argv)>2 and sys.argv[2] in ('A','B') else '')
outdir=sys.argv[-1] if len(sys.argv)>2 else '.'
if ver=='lite': scene=''
VL={'std':'集群标准版','pro':'集群专业版','lite':'集群精简版'}[ver]

# 各版本拓扑
if ver=='pro':
    my,mo,rd,ka,es,mn=".31/.32/.33",".34/.35/.36",".41/.42/.43",".51/.52/.53",".61/.62/.63",".71/.72/.73/.74"
    fl="Flink .81/.82/.83"; myport="3306 / 6446"; mnport="9011-9014"; fport="9001-9004"; co=""
elif ver=='std':
    my,mo,rd,ka,es,mn=".31/.32/.33",".31/.32/.33",".41/.42/.43",".52/.53/.54",".52/.53/.54",".51/.52/.53/.54"
    fl="Flink .61/.62"; myport="3306 / 6446"; mnport="9011-9014"; fport="9001-9004"; co="（与 MongoDB 共置）"
else:  # lite 单节点
    my,mo,rd,ka,es,mn=".31",".31",".31",".51",".51",".51"
    fl="Flink .30"; myport="3306（直连）"; mnport="9011"; fport="9000"; co="（单机共置）"
cluster=(ver!='lite')
# 展开为完整 IP（便于 gen_from_input 用实际 IP 精确替换占位）
def _F(s): return ' / '.join(('192.168.1'+x.strip() if x.strip().startswith('.') else x.strip()) for x in s.split('/'))
my,mo,rd,ka,es,mn=_F(my),_F(mo),_F(rd),_F(ka),_F(es),_F(mn)
fl='Flink '+_F(fl.split(' ',1)[1])

def n(p,ip): return p+' '+ip   # "MySQL 192.168.1.31 / ..."
rows=[]
# MySQL
rows.append(("MySQL",n('MySQL',my),myport,"root","root 初始化密码","<强密码>","mysqld 初始化 / my.cnf","MySQL 部署","= ConfigMap ENV_MYSQL_PASSWORD（约束 C）"+(co if ver=='std' else "")))
rows.append(("MySQL","K8s ConfigMap","—","root","ENV_MYSQL_USERNAME","root","HAP ConfigMap","K8s 微服务",""))
rows.append(("MySQL","K8s ConfigMap","—","root","ENV_MYSQL_PASSWORD","<强密码>","HAP ConfigMap","K8s 微服务","= MySQL root（约束 C）"))
# MongoDB
rows.append(("MongoDB",n('MongoDB',mo),"27017","root","admin 管理账号","<强密码>","mongod 初始化","MongoDB 部署",co))
rows.append(("MongoDB",n('MongoDB',mo),"27017","hap","业务库账号","<强密码>","mongod 初始化","MongoDB 部署","= ConfigMap ENV_MONGODB_URI 内"))
if cluster:
    rows.append(("MongoDB",n('MongoDB',mo),"—","—","副本集 keyFile","<keyFile 内容>","/data/mongodb/keyfile","MongoDB 副本集","三节点完全一致（约束 F）"))
rows.append(("MongoDB","K8s ConfigMap","—","hap","ENV_MONGODB_URI",("mongodb://hap:<强密码>@%s/..."%mo) if cluster else "mongodb://hap:<强密码>@192.168.1.31:27017/...","HAP ConfigMap","K8s 微服务","含账号密码"))
# Redis
if cluster:
    rows.append(("Redis",n('Redis',rd),"6379","—","requirepass 访问密码","<强密码>","/usr/local/redis/redis.conf","Redis 哨兵","三节点一致 = masterauth = Sentinel = ENV（约束 B）"))
    rows.append(("Redis",n('Redis',rd),"6379","—","masterauth 主从密码","<强密码>","/usr/local/redis/redis.conf","Redis 哨兵","= requirepass（约束 B）"))
    rows.append(("Redis",n('Redis',rd),"26379","—","Sentinel auth-pass","<强密码>","/usr/local/redis/sentinel.conf","Redis 哨兵","= requirepass（约束 B）"))
    rows.append(("Redis","K8s ConfigMap","—","—","ENV_REDIS_SENTINEL_PASSWORD","<强密码>","HAP ConfigMap","K8s 微服务","= requirepass（约束 B）"))
    rows.append(("Redis","K8s ConfigMap","—","—","ENV_REDIS_SENTINEL_MASTER","mymaster","HAP ConfigMap","K8s 微服务","Sentinel master 名（非密钥）"))
else:
    rows.append(("Redis",n('Redis',rd),"6379","—","requirepass 访问密码","<强密码>","/usr/local/redis/redis.conf","Redis 部署","单实例 = ConfigMap ENV_REDIS_PASSWORD（约束 B）"))
    rows.append(("Redis","K8s ConfigMap","—","—","ENV_REDIS_PASSWORD","<强密码>","HAP ConfigMap","K8s 微服务","= requirepass（约束 B）"))
# Kafka
rows.append(("Kafka",n('Kafka',ka),"9092 / 2181","（无认证）","—","—","server.properties","Kafka 部署","靠网络隔离，不对外暴露"))
# ES
rows.append(("Elasticsearch",n('ES',es),"9200 / 9300","elastic","x-pack 密码","<强密码>","elasticsearch.yml","Elasticsearch 部署","= ENV_ELASTICSEARCH_PASSWORD（约束 D）"))
rows.append(("Elasticsearch","K8s ConfigMap","—","elastic","ENV_ELASTICSEARCH_PASSWORD","elastic:<强密码>","HAP ConfigMap","K8s 微服务","= ES elastic 密码（约束 D）"))
# MinIO
swnote = ("（场景 %s · "%scene)+("同一 Swarm，Node01 编排）" if scene=='B' else "独立 docker swarm）") if scene else ""
rows.append(("MinIO",n('存储',mn),mnport,"mingdao","MINIO_ROOT_USER","mingdao","minio.yaml"+swnote,"MinIO 部署","= HAP s3 / Flink s3 access-key（约束 A）"))
rows.append(("MinIO",n('存储',mn),mnport,"—","MINIO_ROOT_PASSWORD","<强密码>","minio.yaml","MinIO 部署","= HAP s3 / Flink s3 secret-key（约束 A）"))
rows.append(("HAP s3-config","K8s 微服务","—","mingdao","s3.access-key","mingdao","s3-config.json","MinIO 部署","= MINIO_ROOT_USER（约束 A）"))
rows.append(("HAP s3-config","K8s 微服务","—","—","s3.secret-key","<强密码>","s3-config.json","MinIO 部署","= MINIO_ROOT_PASSWORD（约束 A）"))
# File
rows.append(("File 服务",n('存储',mn),fport,"storage","ENV_ACCESS_KEY_FILE","storage","file 容器环境变量","File 部署","= ConfigMap ENV_FILE_ACCESSKEY（约束 E）"))
rows.append(("File 服务",n('存储',mn),fport,"—","ENV_SECRET_KEY_FILE","<强密码>","file 容器环境变量","File 部署","= ConfigMap ENV_FILE_SECRETKEY（约束 E）"))
rows.append(("File 服务","K8s ConfigMap","—","storage","ENV_FILE_ACCESSKEY / SECRETKEY","storage / <强密码>","HAP ConfigMap","K8s 微服务","= File 容器 AK/SK（约束 E）"))
# Flink
rows.append(("Flink",fl,"—","mingdao","s3.access-key / s3.secret-key","mingdao / <强密码>","flink-conf.yaml","Flink 数据集成","= MINIO_ROOT_USER/PASSWORD（约束 A）"))
# API token
rows.append(("HAP 微服务","K8s ConfigMap","—","—","ENV_API_TOKEN","<高熵随机字符串>","HAP ConfigMap","K8s 微服务","所有微服务副本一致（约束 G）"))
# Keepalived (仅集群版有 VIP)
if cluster:
    rows.append(("Keepalived","Nginx 192.168.1.11 / 192.168.1.12","—","—","auth_pass","HAP-Nginx-Keepalived-Auth","keepalived.conf","Nginx + Keepalived","两节点一致（约束 H）"))
# 入口/平台
rows.append(("K8s 管理器入口","K8s Master 192.168.1.21","38881","（部署设置）","管理器后台口令","（部署时设定）","管理器","K8s 部署","不对公网暴露"))
rows.append(("HAP 超级管理员","hap.domain.com","443","（首次初始化）","平台超级管理员","（首次初始化设定）","Web 初始化","验收","系统最高权限，妥善保管"))
rows.append(("Grafana","监控节点","3000","admin","Grafana 登录","admin（默认，首次改密）","Grafana","监控","首次登录强制改密"))

# 一致性约束
cons=[("A","MinIO 凭据三处一致","minio.yaml MINIO_ROOT_USER/PASSWORD = HAP s3-config.json = Flink flink-conf.yaml 的 s3.access-key/secret-key","mingdao / <强密码>")]
if cluster:
    cons.append(("B","Redis 密码全一致","redis.conf requirepass = masterauth = sentinel.conf auth-pass = ConfigMap ENV_REDIS_SENTINEL_PASSWORD（且 3 节点相同）","<强密码>"))
else:
    cons.append(("B","Redis 密码一致","redis.conf requirepass = ConfigMap ENV_REDIS_PASSWORD（单实例）","<强密码>"))
cons.append(("C","MySQL 密码一致","MySQL root 初始化密码 = ConfigMap ENV_MYSQL_PASSWORD","<强密码>"))
cons.append(("D","ES 密码一致","elasticsearch elastic 密码 = ConfigMap ENV_ELASTICSEARCH_PASSWORD（格式 elastic:<强密码>）","<强密码>"))
cons.append(("E","File 凭据一致","file 容器 ENV_ACCESS_KEY_FILE/ENV_SECRET_KEY_FILE = ConfigMap ENV_FILE_ACCESSKEY/ENV_FILE_SECRETKEY","storage / <强密码>"))
if cluster:
    cons.append(("F","MongoDB keyFile 一致","副本集 3 节点 /data/mongodb/keyfile 内容完全相同（权限 400）","<keyFile 内容>"))
cons.append(("G","API Token 一致","ConfigMap ENV_API_TOKEN 对所有微服务副本一致","<高熵随机字符串>"))
if cluster:
    cons.append(("H","Keepalived 一致","两 Nginx 节点 keepalived.conf auth_pass、virtual_router_id 相同","HAP-Nginx-Keepalived-Auth"))

# ---- 写 xlsx ----
F="Microsoft YaHei"
hdrf=PatternFill("solid",fgColor="1E5BA8"); hdrft=Font(name=F,bold=True,color="FFFFFF",size=10.5)
yfill=PatternFill("solid",fgColor="FFF6CC")
cf=Font(name=F,size=10); tf=Font(name=F,bold=True,size=14,color="1E5BA8"); nf=Font(name=F,size=9,italic=True,color="888888")
thin=Side(style="thin",color="D0D7E5"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
wb=openpyxl.Workbook(); ws=wb.active; ws.title="凭据登记表"
stitle = ("（场景 %s）"%scene) if scene else ""
ws["A1"]="HAP %s · 凭据登记表%s"%(VL,stitle); ws["A1"].font=tf
ws["A2"]="占位符为部署基线示例；实际值列（黄色）由实施工程师部署时填写并加密存档。<强密码> 建议 16+ 位高熵随机。"; ws["A2"].font=nf
cols=["#","组件","节点 / IP","端口","用户名","配置项 / 字段","占位符 / 示例","实际值（待填）","配置文件 / 位置","对应章节","一致性依赖"]
for j,c in enumerate(cols,1):
    cell=ws.cell(4,j,c); cell.fill=hdrf; cell.font=hdrft; cell.alignment=ctr; cell.border=bd
r=5
for i,(comp,node,port,user,key,ph,cfile,chap,dep) in enumerate(rows,1):
    for j,v in enumerate([i,comp,node,port,user,key,ph,"",cfile,chap,dep],1):
        cell=ws.cell(r,j,v); cell.font=cf; cell.border=bd; cell.alignment=ctr if j in (1,4) else lft
        if j==8: cell.fill=yfill
    r+=1
for j,w in enumerate([4,14,18,12,12,24,26,16,22,14,36],1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A5"
w2=wb.create_sheet("一致性约束")
w2["A1"]="凭据一致性约束（部署/迁移时必须保持完全相同）"; w2["A1"].font=tf
for j,c in enumerate(["#","约束","涉及位置","示例值"],1):
    cell=w2.cell(3,j,c); cell.fill=hdrf; cell.font=hdrft; cell.alignment=ctr; cell.border=bd
rr=4
for (a,b,c,d) in cons:
    for j,v in enumerate([a,b,c,d],1):
        cell=w2.cell(rr,j,v); cell.font=cf; cell.border=bd; cell.alignment=ctr if j==1 else lft
    rr+=1
for j,w in enumerate([4,18,70,26],1): w2.column_dimensions[chr(64+j)].width=w
w2.freeze_panes="A4"
fname="HAP%s%s_凭据登记表.xlsx"%(VL, ("_场景"+scene) if scene else "")
os.makedirs(outdir,exist_ok=True); path=os.path.join(outdir,fname)
wb.save(path)
print("生成:",fname,"| 凭据",len(rows),"条 | 约束",len(cons),"组")
